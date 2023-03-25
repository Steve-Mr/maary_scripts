#!/usr/bin/python
# -*-coding:utf-8 -*-
import csv
import sys
from pathlib import Path
import pandas
import datetime
import os
import shutil
from openpyxl import Workbook, load_workbook

# 表格中名字顺序
names_dict = {'杨童旭':'', '马雯雯':'', '杜玉冰':'', '刘首辉':'', '刘郑':'', '孟瑞':'', '陶红辉':'', '高宇恒':'',} 

def process(csv_path):

    # 获取当前日期和七天前日期
    today = datetime.datetime.now()
    intervel = datetime.timedelta(days = 6)
    first_day = (today - intervel).strftime("%Y.%m.%d")
    today = today.strftime("%Y.%m.%d")

    # 打开周报收集表 csv
    with open(csv_path, newline='') as csv_file:
        reader = csv.reader(csv_file)
        for row in reader:
            name = row[2]

            # 第一行 舍弃
            if name == '姓名（必填）':
                continue

            # 填写日期不在本周期内 舍弃
            if datetime.datetime.strptime(row[1].split(" ")[0], "%Y/%m/%d") < datetime.datetime.strptime(first_day, "%Y.%m.%d"):
                continue

            # row[0]：提交者 row[1]：提交时间 row[2]：姓名 row[3]及之后：任务
            row = row[3:]
            index = 1
            report = []
            for cell in row:
                if cell:
                    
                    # 任务中不同项默认使用 | 做分隔，也有使用 ｜ 或不同分隔符的特殊情况
                    fields = cell.split("|")
                    if(len(fields) == 1):
                        fields = cell.split("｜")
                    if(len(fields) == 1):
                        print(cell)

                        # 嗅探未知分隔符
                        sniffer = csv.Sniffer()
                        try:
                            sep = sniffer.sniff(cell).delimiter
                            # print(cell, sep)
                        except:
                            # 嗅探也失败，直接作为单条记录放入表格
                            fields = [cell]
                        else:
                            fields = cell.split(sep)
                    # 插入任务顺序编号
                    fields.insert(0, index)

                    # 相当于空白行进行填充，避免项目不够情况
                    results = ['','','','','']
                    for r_index, field in enumerate(fields):
                        results[r_index] = field
                    index+=1
                    report.append(results)
            names_dict[name] = report

        # 为个人首个任务添加姓名项
        for key, values in names_dict.items():
            if names_dict[key]:
                for index, value in enumerate(values):
                    if(index == 0):
                        value.insert(0, key)
                    else:
                        value.insert(0, '')
                    # 表格中奇怪编号位置
                    value.insert(5,'')
                    
            else:
                # 未填写周报
                values=[[key,'','','','','','']]
            names_dict[key] = values

        # 从模板复制一份，文件名改为今日日期
        shutil.copyfile((Path(csv_path).parent).joinpath("周报.xlsx"), (Path(csv_path).parent).joinpath("{}.xlsx".format(today)))

        # 加载复制的表格
        Workbook = load_workbook((Path(csv_path).parent).joinpath("{}.xlsx".format(today)))
        worksheet = Workbook.active

        # 表格首行内容修改
        worksheet.cell(row=1, column=1).value = '任务总结 时间：{}-{}'.format(first_day, today)

        # 表格内容从第三行开始
        row_index = 3
        for key in names_dict:
            # 如果某人提交多条记录，则需要将名字部分合并单元格
            if len(names_dict[key]) != 1:
                # 合并前清空内容
                for i in range(row_index, row_index+len(names_dict[key])-1):
                    worksheet.cell(i,1).value = ''
                worksheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index+len(names_dict[key])-1, end_column=1)
            # 填写单元格内容
            for value in names_dict[key]:
                for col_index, field in enumerate(value):
                    # 被合并的单元格不再填写内容
                    if col_index == 0 and field == "":
                        continue
                    worksheet.cell(row=row_index, column=col_index+1).value = field
                row_index += 1
        # 清空其他单元格
        if row_index < worksheet.max_row:
            for i in range(row_index, worksheet.max_row):
                for j in range(1, worksheet.max_column):
                    worksheet.cell(i, j).value = ''

        # 保存文件
        Workbook.save((Path(csv_path).parent).joinpath("{}.xlsx".format(today)))    

if __name__ == "__main__":
    print(sys.argv)
    # process('/home/maary/Downloads/周报（220828）（收集结果）-周报（220828）（收集结果）(1).csv')
    process(sys.argv[1])
    print('finished')
